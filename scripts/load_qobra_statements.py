#!/usr/bin/env python3
"""Parse Qobra manual statement exports (qobra/*.xlsx) and load them into Supabase.

One-time / re-runnable import: since the Qobra API's reporting endpoints are not
accessible with the current API key, monthly statements are exported manually
from the Qobra UI as .xlsx and dropped in the qobra/ folder. This script parses
them and upserts into three Supabase tables:
  - qobra_rep_performance   (quota vs. achieved, one row per rep per month)
  - qobra_bonus_line_items  (per-account bonus detail)
  - qobra_opportunities     (per-opportunity detail feeding the bonus)
"""
import glob
import json
import os
import re
import urllib.request
from datetime import date, datetime

import openpyxl

QOBRA_DIR = os.path.join(os.path.dirname(__file__), "..", "qobra")
ENV_PATH = os.path.join(os.path.dirname(__file__), "..", ".env")


def load_env():
    env = {}
    with open(ENV_PATH) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k] = v
    return env


def sql_str(v):
    if v is None:
        return "NULL"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, (datetime, date)):
        return f"'{v.isoformat()}'"
    return "'" + str(v).replace("'", "''") + "'"


def run_sql(query, access_token, project_ref):
    req = urllib.request.Request(
        f"https://api.supabase.com/v1/projects/{project_ref}/database/query",
        data=json.dumps({"query": query}).encode(),
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
            "User-Agent": "curl/8.7.1",
        },
        method="POST",
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


SCHEMA_SQL = """
create table if not exists qobra_rep_performance (
    performance_name text primary key,
    owner_email text not null,
    owner_name text not null,
    owner_role text,
    owner_team text,
    performance_month date not null,
    monthly_quota numeric,
    onboarded_amnr numeric,
    signed_emnr numeric,
    currency text,
    created_date date,
    loaded_at timestamptz default now()
);

create table if not exists qobra_bonus_line_items (
    unique_key text primary key,
    owner_email text not null,
    account_name text,
    product text,
    source text,
    bonus_start_date date,
    bonus_end_date date,
    net_revenue numeric,
    adj_net_revenue numeric,
    currency text,
    sales_factor numeric,
    trial_offer_factor numeric,
    country_factor numeric,
    seasonal_factor numeric,
    adj_nr_new_acquisition numeric,
    adj_nr_upsell numeric,
    adj_nr_group_expansion numeric,
    loaded_at timestamptz default now()
);

create table if not exists qobra_opportunities (
    opportunity_name text,
    owner_email text not null,
    owner_name text,
    created_date date,
    close_date date,
    estimated_monthly_nr numeric,
    nr_first_28_days numeric,
    currency text,
    opportunity_source text,
    related_products text,
    opportunity_type text,
    loaded_at timestamptz default now(),
    primary key (opportunity_name, owner_email)
);

create or replace view qobra_rep_attainment as
select
    owner_email,
    owner_name,
    owner_team,
    performance_month,
    monthly_quota,
    onboarded_amnr,
    currency,
    round(onboarded_amnr / nullif(monthly_quota, 0) * 100, 1) as attainment_pct
from qobra_rep_performance;
"""


def parse_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    ws_c = wb["Individual Bonus (Monthly) - C"]
    c_header = [c.value for c in ws_c[1]]
    c_row = [c.value for c in ws_c[2]]
    c = dict(zip(c_header, c_row))

    owner_email = c["Owner Email"]

    performance = {
        "performance_name": c["User Performance Name"],
        "owner_email": owner_email,
        "owner_name": c["Owner Name"],
        "owner_role": c["Owner Internal Role"],
        "owner_team": c["Owner Team"],
        "performance_month": c["Performance Month"],
        "monthly_quota": c["Monthly Quota"],
        "onboarded_amnr": c["Onboarded AMNR"],
        "signed_emnr": c["Signed EMNR"],
        "currency": c["Monthly Quota Currency"],
        "created_date": c["Created Date"],
    }

    ws_a = wb["Individual Bonus (Monthly) - A"]
    a_header = [c.value for c in ws_a[1]]
    bonus_items = []
    for row in ws_a.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        r = dict(zip(a_header, row))
        bonus_items.append({
            "unique_key": r["Unique Key"],
            "owner_email": owner_email,
            "account_name": r["Account Name"],
            "product": r["Product"],
            "source": r["Source"],
            "bonus_start_date": r["Bonus Start Date"],
            "bonus_end_date": r["Bonus End Date"],
            "net_revenue": r["Net Revenue"],
            "adj_net_revenue": r["Adj Net Revenue"],
            "currency": r["Net Revenue Currency"],
            "sales_factor": r["Sales Factor"],
            "trial_offer_factor": r["Trial Offer Factor"],
            "country_factor": r["Country Factor"],
            "seasonal_factor": r["Seasonal Factor"],
            "adj_nr_new_acquisition": r["Adj NR New Acquisition"],
            "adj_nr_upsell": r["Adj. NR Upsell"],
            "adj_nr_group_expansion": r["Adj. NR Group Expansion"],
        })

    ws_b = wb["Individual Bonus (Monthly) - B"]
    b_header = [c.value for c in ws_b[1]]
    opportunities = []
    for row in ws_b.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        r = dict(zip(b_header, row))
        opportunities.append({
            "opportunity_name": r["Name"],
            "owner_email": r["Owner Email"],
            "owner_name": r["Owner Name"],
            "created_date": r["Created Date"],
            "close_date": r["Close Date"],
            "estimated_monthly_nr": r["Estimated Monthly Net Revenue"],
            "nr_first_28_days": r["sunday Net Revenue - First 28 Days"],
            "currency": r["Estimated Monthly Net Revenue Currency"],
            "opportunity_source": r["Opportunity Source"],
            "related_products": r["Related Products"],
            "opportunity_type": r["Opportunity Type"],
        })

    return performance, bonus_items, opportunities


def build_upsert(table, rows, conflict_cols):
    if not rows:
        return ""
    cols = list(rows[0].keys())
    values_sql = ",\n".join(
        "(" + ", ".join(sql_str(row[col]) for col in cols) + ")" for row in rows
    )
    update_sql = ", ".join(f"{col} = excluded.{col}" for col in cols if col not in conflict_cols)
    conflict_sql = ", ".join(conflict_cols)
    return (
        f"insert into {table} ({', '.join(cols)}) values\n{values_sql}\n"
        f"on conflict ({conflict_sql}) do update set {update_sql};"
    )


def main():
    env = load_env()
    access_token = env["SUPABASE_ACCESS_TOKEN"]
    project_ref = "jvmebywlusemucgbomgl"

    print("Creating schema...")
    run_sql(SCHEMA_SQL, access_token, project_ref)

    all_perf, all_items, all_opps = [], [], []
    for path in sorted(glob.glob(os.path.join(QOBRA_DIR, "*.xlsx"))):
        perf, items, opps = parse_file(path)
        all_perf.append(perf)
        all_items.extend(items)
        all_opps.extend(opps)
        print(f"Parsed {os.path.basename(path)}: 1 perf row, {len(items)} bonus items, {len(opps)} opportunities")

    sql = "\n".join([
        build_upsert("qobra_rep_performance", all_perf, ["performance_name"]),
        build_upsert("qobra_bonus_line_items", all_items, ["unique_key"]),
        build_upsert("qobra_opportunities", all_opps, ["opportunity_name", "owner_email"]),
    ])

    print("Loading data into Supabase...")
    run_sql(sql, access_token, project_ref)
    print("Done.")


if __name__ == "__main__":
    main()
